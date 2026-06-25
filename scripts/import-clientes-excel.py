#!/usr/bin/env python3
"""
Importa / sincroniza clientes do Excel «Base de Dados.xlsx» para Supabase.

Colunas esperadas (folha «Base de Dados - Com Localidades»):
  Clientes, NIF, Email, Morada, Código postal, Localidade, Plus Code+, Zona / Rota

Uso:
  python scripts/import-clientes-excel.py
      # pré-visualização (dry-run) — precisa SUPABASE_SERVICE_ROLE_KEY no ambiente

  python scripts/import-clientes-excel.py --sql-out pwa/supabase/migrations/013_sync_clientes_excel.sql
      # gera SQL UPDATE por NIF (sem credenciais Supabase)

  python scripts/import-clientes-excel.py --apply
      # aplica alterações via API REST (service role)

Variáveis de ambiente:
  SUPABASE_URL              (default: projeto ManuSilva)
  SUPABASE_SERVICE_ROLE_KEY obrigatória para --dry-run / --apply
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DEFAULT_XLSX = Path(r"\\DI\Manusilva Escritório\Joana\Informações\Base de Dados.xlsx")
DEFAULT_SHEET = "Base de Dados - Com Localidades"
DEFAULT_SUPABASE_URL = "https://zhfbezrevosmbmcbyskw.supabase.co"

COLUMNS = (
    "nome",
    "nif",
    "email",
    "morada",
    "codigo_postal",
    "localidade",
    "plus_code",
    "zona_rota",
)


def norm_nif(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = re.sub(r"\D", "", str(value))
    return s


def norm_text(value) -> str | None:
    if value is None:
        return None
    s = str(value).replace("\xa0", " ").strip()
    return s or None


def norm_nome(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def load_excel_rows(path: Path, sheet_name: str) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"Excel não encontrado: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Folha «{sheet_name}» não encontrada. Folhas: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or not raw[0]:
            continue
        values = list(raw[:8]) + [None] * (8 - len(raw[:8]))
        item = dict(zip(COLUMNS, values))
        item["nif"] = norm_nif(item["nif"])
        for key in ("email", "morada", "codigo_postal", "localidade", "plus_code", "zona_rota"):
            item[key] = norm_text(item[key])
        item["nome"] = norm_text(item["nome"]) or ""
        rows.append(item)

    wb.close()
    return rows


def fetch_supabase_clients(url: str, service_key: str) -> list[dict]:
    endpoint = (
        f"{url.rstrip('/')}/rest/v1/clientes"
        "?select=id,nome_empresa,nif,email,morada,codigo_postal,localidade,plus_code,zona_rota"
        "&order=id.asc"
    )
    req = urllib.request.Request(
        endpoint,
        headers={
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
        },
    )
    with urllib.request.urlopen(req, timeout=120) as res:
        return json.loads(res.read().decode("utf-8"))


def patch_supabase_client(url: str, service_key: str, client_id: int, patch: dict) -> None:
    endpoint = f"{url.rstrip('/')}/rest/v1/clientes?id=eq.{client_id}"
    body = json.dumps(patch).encode("utf-8")
    req = urllib.request.Request(
        endpoint,
        data=body,
        method="PATCH",
        headers={
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
    )
    with urllib.request.urlopen(req, timeout=60):
        return


def build_indexes(clients: list[dict]) -> tuple[dict[str, dict], dict[str, dict]]:
    by_nif: dict[str, dict] = {}
    by_nome: dict[str, dict] = {}
    for c in clients:
        nif = norm_nif(c.get("nif"))
        if nif and nif not in by_nif:
            by_nif[nif] = c
        nome = norm_nome(c.get("nome_empresa"))
        if nome and nome not in by_nome:
            by_nome[nome] = c
    return by_nif, by_nome


def match_excel_to_db(excel_rows: list[dict], db_clients: list[dict]) -> tuple[list[dict], list[dict]]:
    by_nif, by_nome = build_indexes(db_clients)
    matched: list[dict] = []
    unmatched: list[dict] = []

    for row in excel_rows:
        db = None
        if row["nif"]:
            db = by_nif.get(row["nif"])
        if not db and row["nome"]:
            db = by_nome.get(norm_nome(row["nome"]))
        if not db:
            unmatched.append(row)
            continue

        patch = {}
        mapping = {
            "email": "email",
            "morada": "morada",
            "codigo_postal": "codigo_postal",
            "localidade": "localidade",
            "plus_code": "plus_code",
            "zona_rota": "zona_rota",
        }
        for excel_key, db_key in mapping.items():
            new_val = row.get(excel_key)
            if new_val is None:
                continue
            old_val = norm_text(db.get(db_key))
            if old_val != new_val:
                patch[db_key] = new_val

        if patch:
            matched.append(
                {
                    "id": db["id"],
                    "nome_excel": row["nome"],
                    "nome_db": db.get("nome_empresa"),
                    "nif": row["nif"] or norm_nif(db.get("nif")),
                    "patch": patch,
                }
            )

    return matched, unmatched


def generate_sql(_matched: list[dict], excel_rows: list[dict]) -> str:
    """Gera SQL com VALUES inline (compatível com Supabase SQL Editor — sem TEMP TABLE)."""
    value_rows: list[str] = []
    for row in excel_rows:
        nif_digits = row.get("nif") or ""
        cols = [
            sql_literal(nif_digits),
            sql_literal(row.get("nome")),
            sql_literal(row.get("email")),
            sql_literal(row.get("morada")),
            sql_literal(row.get("codigo_postal")),
            sql_literal(row.get("localidade")),
            sql_literal(row.get("plus_code")),
            sql_literal(row.get("zona_rota")),
        ]
        value_rows.append(f"    ({', '.join(cols)})")

    lines = [
        "-- Gerado por scripts/import-clientes-excel.py",
        f"-- {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        "-- Executar no Supabase → SQL Editor (após 012_plus_code_zona_rota.sql)",
        "-- Correspondência: NIF só dígitos OU nome (trim, case-insensitive).",
        "-- Nota: colar e executar o bloco completo de uma vez.",
        "",
        "UPDATE public.clientes AS c",
        "SET",
        "  email = COALESCE(s.email, c.email),",
        "  morada = COALESCE(s.morada, c.morada),",
        "  codigo_postal = COALESCE(s.codigo_postal, c.codigo_postal),",
        "  localidade = COALESCE(s.localidade, c.localidade),",
        "  plus_code = COALESCE(s.plus_code, c.plus_code),",
        "  zona_rota = COALESCE(s.zona_rota, c.zona_rota)",
        "FROM (",
        "  VALUES",
        ",\n".join(value_rows),
        ") AS s(nif_digits, nome, email, morada, codigo_postal, localidade, plus_code, zona_rota)",
        "WHERE (",
        "  s.nif_digits <> ''",
        "  AND regexp_replace(COALESCE(c.nif::text, ''), '[^0-9]', '', 'g') = s.nif_digits",
        ") OR (",
        "  lower(trim(c.nome_empresa)) = lower(trim(s.nome))",
        ");",
        "",
        "-- Verificação (deve ser > 0 após importação)",
        "SELECT",
        "  COUNT(*) FILTER (WHERE plus_code IS NOT NULL AND btrim(plus_code) <> '') AS com_plus_code,",
        "  COUNT(*) FILTER (WHERE zona_rota IS NOT NULL AND btrim(zona_rota) <> '') AS com_zona_rota",
        "FROM public.clientes;",
        "",
    ]
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Sincronizar clientes Excel → Supabase")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Caminho do Excel")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Nome da folha")
    parser.add_argument("--sql-out", type=Path, help="Gerar ficheiro SQL com UPDATEs")
    parser.add_argument("--apply", action="store_true", help="Aplicar PATCH na Supabase")
    parser.add_argument("--dry-run", action="store_true", help="Mostrar alterações sem gravar")
    args = parser.parse_args()

    dry_run = args.dry_run or (not args.apply and not args.sql_out)

    try:
        excel_rows = load_excel_rows(args.xlsx, args.sheet)
    except (FileNotFoundError, ValueError) as err:
        print(f"Erro: {err}", file=sys.stderr)
        return 1

    print(f"Excel: {len(excel_rows)} clientes lidos de «{args.sheet}»")

    supabase_url = os.environ.get("SUPABASE_URL", DEFAULT_SUPABASE_URL).strip()
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()

    matched: list[dict] = []
    unmatched: list[dict] = []

    if service_key:
        try:
            db_clients = fetch_supabase_clients(supabase_url, service_key)
            print(f"Supabase: {len(db_clients)} clientes na base de dados")
            matched, unmatched = match_excel_to_db(excel_rows, db_clients)
        except urllib.error.HTTPError as err:
            body = err.read().decode("utf-8", errors="replace")
            print(f"Erro HTTP Supabase ({err.code}): {body}", file=sys.stderr)
            if not args.sql_out:
                return 1
        except urllib.error.URLError as err:
            print(f"Erro de rede Supabase: {err}", file=sys.stderr)
            if not args.sql_out:
                return 1
    else:
        print("SUPABASE_SERVICE_ROLE_KEY não definida — só é possível gerar SQL (--sql-out).")

    if matched:
        print(f"Correspondências com alterações: {len(matched)}")
        for item in matched[:8]:
            fields = ", ".join(item["patch"].keys())
            print(f"  • [{item['id']}] {item['nome_excel']} → {fields}")
        if len(matched) > 8:
            print(f"  … e mais {len(matched) - 8}")
    elif service_key:
        print("Nenhuma alteração pendente (dados já sincronizados ou sem match).")

    if unmatched:
        print(f"\nSem correspondência na DB ({len(unmatched)}):")
        for row in unmatched[:10]:
            print(f"  • {row['nome']} (NIF {row['nif'] or '—'})")
        if len(unmatched) > 10:
            print(f"  … e mais {len(unmatched) - 10}")

    if args.sql_out:
        sql = generate_sql(matched, excel_rows)
        args.sql_out.parent.mkdir(parents=True, exist_ok=True)
        args.sql_out.write_text(sql, encoding="utf-8")
        print(f"\nSQL gravado em: {args.sql_out}")

    if args.apply:
        if not service_key:
            print("SUPABASE_SERVICE_ROLE_KEY em falta.", file=sys.stderr)
            return 1
        if not matched:
            print("Nada para aplicar.")
            return 0
        for item in matched:
            patch_supabase_client(supabase_url, service_key, item["id"], item["patch"])
        print(f"Aplicadas {len(matched)} atualizações.")
    elif dry_run and service_key and matched:
        print("\n(dry-run — usa --apply para gravar ou --sql-out para gerar SQL)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
