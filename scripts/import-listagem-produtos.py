#!/usr/bin/env python3
"""
Importa «Listagem Produtos.xlsx» → pwa/data/catalogo-produtos.json

Colunas esperadas (cabeçalho na 1.ª linha):
  Tipo | Código | Descrição / Nome | Uni. | Preço de venda

Uso:
  python scripts/import-listagem-produtos.py
  python scripts/import-listagem-produtos.py "data/Listagem-Produtos.xlsx"
  python scripts/import-listagem-produtos.py "\\\\DI\\Manusilva Escritório\\Joana\\Listagem Produtos.xlsx"
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SRC = ROOT / "data" / "Listagem-Produtos.xlsx"
OUT = ROOT / "pwa" / "data" / "catalogo-produtos.json"

HEADER_MAP = {
    "tipo": ("tipo",),
    "codigo": ("código", "codigo", "code", "ref", "referência", "referencia"),
    "descricao": (
        "descrição / nome",
        "descricao / nome",
        "descrição",
        "descricao",
        "nome",
        "descrição/nome",
        "descricao/nome",
    ),
    "unidade": ("uni.", "uni", "unidade", "un", "ud"),
    "preco_venda": (
        "preço de venda",
        "preco de venda",
        "preço venda",
        "preco venda",
        "pvp",
        "preço",
        "preco",
        "preço de venda (€)",
    ),
}


def norm_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ")
    return re.sub(r"\s+", " ", text)


def parse_price(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        return n if n >= 0 else None
    text = str(value).strip().replace("\u00a0", " ").replace("€", "").strip()
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".") if "," in text and "." in text else text.replace(",", ".")
    try:
        n = float(text)
    except ValueError:
        return None
    return n if n >= 0 else None


def resolve_columns(headers: list[object]) -> dict[str, int]:
    normalized = [norm_header(h) for h in headers]
    resolved: dict[str, int] = {}
    for key, aliases in HEADER_MAP.items():
        for idx, header in enumerate(normalized):
            if header in aliases:
                resolved[key] = idx
                break
    missing = [k for k in HEADER_MAP if k not in resolved]
    if missing:
        raise SystemExit(
            f"Colunas em falta no Excel: {', '.join(missing)}\n"
            f"Cabeçalhos encontrados: {headers}"
        )
    return resolved


def cell(row: tuple, index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    if value is None:
        return ""
    return str(value).strip()


def import_workbook(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Instale openpyxl: pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise SystemExit("Folha vazia.")

    cols = resolve_columns(list(rows[0]))
    items: list[dict] = []
    seen: set[str] = set()

    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        descricao = cell(row, cols["descricao"])
        codigo = cell(row, cols["codigo"])
        if not descricao and not codigo:
            continue
        preco = parse_price(row[cols["preco_venda"]] if cols["preco_venda"] < len(row) else None)
        item = {
            "tipo": cell(row, cols["tipo"]) or "—",
            "codigo": codigo,
            "descricao": descricao or codigo,
            "unidade": cell(row, cols["unidade"]) or "un",
            "precoVenda": preco,
        }
        dedupe_key = f"{item['codigo'].lower()}|{item['descricao'].lower()}"
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        items.append(item)

    items.sort(key=lambda x: (x["descricao"].lower(), x["codigo"].lower()))
    return items


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SRC
    if not src.is_file():
        raise SystemExit(
            f"Ficheiro não encontrado: {src}\n"
            "Copie o Excel para data/Listagem-Produtos.xlsx ou passe o caminho como argumento."
        )

    items = import_workbook(src)
    payload = {
        "version": 1,
        "updatedAt": datetime.now(timezone.utc).isoformat(),
        "source": src.name,
        "itemCount": len(items),
        "items": items,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Importados {len(items)} artigos -> {OUT}")


if __name__ == "__main__":
    main()
